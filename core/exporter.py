import json
from datetime import datetime, date
from pathlib import Path
from typing import List, Dict, Any, Optional
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from core.config import OUTPUT_DIR

def format_excel_sheet(ws, title_color="1E293B"):
    """Aplica formatação premium em uma aba do Excel (openpyxl)."""
    header_fill = PatternFill(start_color=title_color, end_color=title_color, fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=10)
    alt_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    # Formatar cabeçalho
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[1].height = 28

    # Formatar dados
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        ws.row_dimensions[row_idx].height = 20
        is_even = (row_idx % 2 == 0)
        
        for cell in row:
            cell.font = data_font
            cell.border = thin_border
            if is_even:
                cell.fill = alt_fill
                
            # Formatação de Moeda
            if isinstance(cell.value, (int, float)) and "valor" in str(ws.cell(row=1, column=cell.column).value).lower():
                cell.number_format = 'R$ #,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # Ajustar largura automática das colunas
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

def export_offers_data(records: List[Dict[str, Any]], custom_tag: str = "") -> Dict[str, Any]:
    """
    Gera arquivos Excel (.xlsx), CSV e JSON estilizados na pasta output/.
    Retorna caminhos dos arquivos gerados e estatísticas resumidas.
    """
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    tag = f"_{custom_tag}" if custom_tag else ""
    
    base_name = f"ofertas_{timestamp}{tag}"
    excel_path = OUTPUT_DIR / f"{base_name}.xlsx"
    json_path = OUTPUT_DIR / f"{base_name}.json"
    csv_path = OUTPUT_DIR / f"{base_name}.csv"
    latest_json_path = OUTPUT_DIR / "latest_results.json"

    if not records:
        empty_data = {
            "timestamp": timestamp,
            "total_itens": 0,
            "supermercados": [],
            "items": []
        }
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(empty_data, f, indent=2, ensure_ascii=False)
        with open(latest_json_path, "w", encoding="utf-8") as f:
            json.dump(empty_data, f, indent=2, ensure_ascii=False)
        return {
            "excel_file": "",
            "json_file": str(json_path),
            "total_items": 0
        }

    df = pd.DataFrame(records)
    
    # Ordenar por Categoria e Valor
    if "categoria" in df.columns and "valor" in df.columns:
        df = df.sort_values(by=["categoria", "valor"], ascending=[True, True])

    # 1. Salvar JSON completo
    export_payload = {
        "timestamp": timestamp,
        "total_itens": len(records),
        "supermercados": sorted(list(set(df["supermercado"].tolist()))) if "supermercado" in df.columns else [],
        "items": records
    }
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(export_payload, f, indent=2, ensure_ascii=False)
    with open(latest_json_path, "w", encoding="utf-8") as f:
        json.dump(export_payload, f, indent=2, ensure_ascii=False)

    # 2. Salvar CSV
    latest_csv_path = OUTPUT_DIR / "latest_results.csv"
    df.to_csv(csv_path, index=False, encoding="utf-8-sig")
    df.to_csv(latest_csv_path, index=False, encoding="utf-8-sig")

    # 3. Gerar Excel com Múltiplas Abas Estilizadas
    latest_excel_path = OUTPUT_DIR / "latest_results.xlsx"
    for current_excel_target in [excel_path, latest_excel_path]:
        with pd.ExcelWriter(current_excel_target, engine="openpyxl") as writer:
            # Aba Principal: Todas as Ofertas Padronizadas
            df_export = df.copy()
            colunas_ordem = [c for c in [
                "supermercado", "categoria", "produto_padronizado", "marca", "embalagem",
                "item", "valor", "data_postagem", "link", "post_url"
            ] if c in df_export.columns]
            df_export[colunas_ordem].to_excel(writer, sheet_name="Todas as Ofertas", index=False)
            
            # Aba 2: Resumo por Supermercado
            if "supermercado" in df.columns and "valor" in df.columns:
                resumo_mercados = df.groupby("supermercado")["valor"].agg(
                    Qtd_Itens="count",
                    Preco_Medio="mean",
                    Preco_Minimo="min",
                    Preco_Maximo="max"
                ).reset_index()
                resumo_mercados.to_excel(writer, sheet_name="Por Supermercado", index=False)

            # Aba 3: Resumo por Categoria
            if "categoria" in df.columns and "valor" in df.columns:
                resumo_cat = df.groupby("categoria")["valor"].agg(
                    Qtd_Itens="count",
                    Preco_Medio="mean",
                    Preco_Minimo="min"
                ).reset_index()
                resumo_cat.to_excel(writer, sheet_name="Por Categoria", index=False)

        # Aplicar estilização visual nas abas do Excel
        from openpyxl import load_workbook
        wb = load_workbook(current_excel_target)
        for sheet_name in wb.sheetnames:
            format_excel_sheet(wb[sheet_name])
        wb.save(current_excel_target)

    return {
        "excel_file": str(excel_path.name),
        "json_file": str(json_path.name),
        "csv_file": str(csv_path.name),
        "total_items": len(records),
        "timestamp": timestamp
    }

def export_comparison_excel(comparison_records: List[Dict[str, Any]]) -> str:
    """
    Gera um relatório Excel comparativo sofisticado com Matriz Produto x Supermercado
    destacando automaticamente o menor preço em verde.
    """
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    target_path = OUTPUT_DIR / f"comparativo_precos_{timestamp}.xlsx"

    # Prepara linhas para a tabela comparativa
    flat_rows = []
    all_supermarkets = set()

    for item in comparison_records:
        row_dict = {
            "Produto Padronizado": item.get("produto_padronizado", ""),
            "Categoria": item.get("categoria", ""),
            "Marca": item.get("marca", ""),
            "Menor Preço": item.get("menor_preco", 0.0),
            "Maior Preço": item.get("maior_preco", 0.0),
            "Preço Médio": item.get("preco_medio", 0.0),
            "Economia (R$)": item.get("economia_reais", 0.0),
            "Economia (%)": f"{item.get('economia_pct', 0.0)}%",
            "Supermercado Mais Barato": item.get("supermercado_mais_barato", "")
        }
        # Adiciona colunas para cada mercado
        for m in item.get("mercados", []):
            mkt_name = m["supermercado"]
            all_supermarkets.add(mkt_name)
            row_dict[f"Preço: {mkt_name}"] = m["valor"]

        flat_rows.append(row_dict)

    df_comp = pd.DataFrame(flat_rows) if flat_rows else pd.DataFrame(columns=["Produto Padronizado", "Menor Preço"])

    with pd.ExcelWriter(target_path, engine="openpyxl") as writer:
        df_comp.to_excel(writer, sheet_name="Comparativo de Preços", index=False)

    from openpyxl import load_workbook
    wb = load_workbook(target_path)
    ws = wb["Comparativo de Preços"]
    format_excel_sheet(ws, title_color="0F766E")

    # Destaque de Menor Preço em Verde Suave
    green_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    green_font = Font(name="Segoe UI", size=10, bold=True, color="166534")

    # Localiza coluna de Menor Preço e colunas de supermercados
    header_cols = {cell.value: cell.column for cell in ws[1]}
    menor_preco_col = header_cols.get("Menor Preço")

    if menor_preco_col:
        market_cols = [col_idx for col_name, col_idx in header_cols.items() if str(col_name).startswith("Preço: ")]
        for row_idx in range(2, ws.max_row + 1):
            menor_val = ws.cell(row=row_idx, column=menor_preco_col).value
            if isinstance(menor_val, (int, float)):
                for m_col in market_cols:
                    cell = ws.cell(row=row_idx, column=m_col)
                    if cell.value == menor_val:
                        cell.fill = green_fill
                        cell.font = green_font

    wb.save(target_path)
    return target_path.name

def get_latest_results() -> Optional[Dict[str, Any]]:
    """Carrega os resultados da última raspagem salva."""
    latest_path = OUTPUT_DIR / "latest_results.json"
    if latest_path.exists():
        try:
            with open(latest_path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return None
    return None

def list_history_runs() -> List[Dict[str, Any]]:
    """Lista execuções anteriores na pasta output/."""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    history = []
    for file in sorted(OUTPUT_DIR.glob("ofertas_*.xlsx"), reverse=True):
        history.append({
            "filename": file.name,
            "size_kb": round(file.stat().st_size / 1024, 1),
            "date": datetime.fromtimestamp(file.stat().st_mtime).strftime("%d/%m/%Y %H:%M:%S")
        })
    return history
